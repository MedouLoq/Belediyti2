from django.shortcuts import render

# Create your views here.
def admin_home(request):
    return render(request,'muni_admin/base.html')

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.utils.translation import gettext as _
from django.db.models import Q, Count
from Citoyen.models import Problem,Category #Import model is core class value dispatcher name

ITEMS_PER_PAGE = 10 #You can set list value

@login_required #if value name login check view dispatcher access  for not name access
def problem_list(request):
    user_municipality = request.user.admin_profile.municipality #user by session name in model with one filter to view list is user

    search_term = request.GET.get('search', '')#dispatch all filter with names
    status_filter = request.GET.get('status', '')  #dispatch all class
    category_filter = request.GET.get('category', '')

    problems = Problem.objects.filter(municipality=user_municipality)#Load models by municipality code list

    # Apply filters
    if search_term:
         problems = problems.filter(Q(description__icontains=search_term) | #number dispatcher user in page if
                                  Q(location__icontains=search_term))# check values text

        #If used this list, be check user if it's a superadmin
    if status_filter:
        problems = problems.filter(status=status_filter)

    if category_filter:
        problems = problems.filter(category_id=category_filter)


    # Pagination use in load last line
    paginator = Paginator(problems, ITEMS_PER_PAGE)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)


    # all new name to load template list by name in order
    context = {
        'problems': page_obj,
        'search_term': search_term,  # value  filter text dispatcher number

       'status_filter': status_filter, #dispatcher name chart

        'category_filter': category_filter,# if used set code values
        'categories': Category.objects.all(),#select option id name function list
        'status_choices': Problem.STATUS_CHOICES,  #model list for select dropdownMenu  class options value
         'navName': 'problems',
         #all names and dispatcher for template list class, and dispatcher function list
    }

    return render(request, 'muni_admin/problem_list.html', context)
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Count, Avg, F, ExpressionWrapper, fields, Q
from django.core import serializers
from Citoyen.models import Problem, Complaint, Category, StatusLog
import datetime
from collections import defaultdict
import json

@login_required
def dashboard(request):
    municipality = request.user.admin_profile.municipality
    import datetime
    # Get period parameter (default: 7 days)
    days_period = int(request.GET.get('days', 7))
    today = timezone.now().date()
    period_start = today - datetime.timedelta(days=days_period)
    previous_period_start = period_start - datetime.timedelta(days=days_period)
    
    # Basic counts
    total_problems = Problem.objects.filter(municipality=municipality).count()
    new_problems_last_period = Problem.objects.filter(
        municipality=municipality, 
        created_at__date__gte=period_start
    ).count()
    
    # Compare with previous period
    problems_previous_period = Problem.objects.filter(
        municipality=municipality,
        created_at__date__gte=previous_period_start,
        created_at__date__lt=period_start
    ).count()
    
    # Calculate percent change
    problem_percent_change = 0
    if problems_previous_period > 0:
        problem_percent_change = round(((new_problems_last_period - problems_previous_period) / problems_previous_period) * 100)
    
    # Same for complaints
    total_complaints = Complaint.objects.filter(municipality=municipality).count()
    new_complaints_last_period = Complaint.objects.filter(
        municipality=municipality, 
        created_at__date__gte=period_start
    ).count()
    
    complaints_previous_period = Complaint.objects.filter(
        municipality=municipality,
        created_at__date__gte=previous_period_start,
        created_at__date__lt=period_start
    ).count()
    
    complaint_percent_change = 0
    if complaints_previous_period > 0:
        complaint_percent_change = round(((new_complaints_last_period - complaints_previous_period) / complaints_previous_period) * 100)
    
    # New problems percent change
    new_problem_percent_change = 0
    if problems_previous_period > 0:
        new_problem_percent_change = round(((new_problems_last_period - problems_previous_period) / problems_previous_period) * 100)
    
    # New complaints percent change
    new_complaint_percent_change = 0
    if complaints_previous_period > 0:
        new_complaint_percent_change = round(((new_complaints_last_period - complaints_previous_period) / complaints_previous_period) * 100)
    
    # Status breakdown for problems
    problem_status_data = Problem.objects.filter(
        municipality=municipality
    ).values('status').annotate(count=Count('status'))
    
    # Status breakdown for complaints
    complaint_status_data = Complaint.objects.filter(
        municipality=municipality
    ).values('status').annotate(count=Count('status'))
    
    # Top problem categories
    problem_categories_data = Problem.objects.filter(
        municipality=municipality
    ).values('category__name').annotate(
        count=Count('category')
    ).order_by('-count')[:5]
    
    # Problem and complaint time series data (for the charts)
    # Get data for last X days
    time_series_days = 90  # Show data for last 90 days
    time_series_start = today - datetime.timedelta(days=time_series_days)
    
    # Generate date range
    date_range = []
    current_date = time_series_start
    while current_date <= today:
        date_range.append(current_date)
        current_date += datetime.timedelta(days=1)
    
    import datetime
    from collections import defaultdict
    
    # Problems by day
    problems_by_day = Problem.objects.filter(
        municipality=municipality,
        created_at__date__gte=time_series_start
    ).annotate(
        created_day=ExpressionWrapper(
            F('created_at__date'),
            output_field=fields.DateField()
        )
    ).values('created_day').annotate(
        count=Count('id')
    ).order_by('created_day')
    
    # Convert to dictionary for easier lookup
    problems_data = defaultdict(int)
    for item in problems_by_day:
        day_str = item['created_day']  # Changed from 'day' to 'created_day'
        if isinstance(day_str, str):
            # Parse the string to datetime first
            day_obj = datetime.datetime.strptime(day_str, '%Y-%m-%d')  # Added datetime prefix
            formatted_day = day_obj.strftime('%Y-%m-%d')
        else:
            formatted_day = day_str.strftime('%Y-%m-%d')
        problems_data[formatted_day] = item['count']
    
    # Complaints by day
    complaints_by_day = Complaint.objects.filter(
        municipality=municipality, 
        created_at__date__gte=time_series_start
    ).annotate(
        created_day=ExpressionWrapper(
            F('created_at__date'), 
            output_field=fields.DateField()
        )
    ).values('created_day').annotate(
        count=Count('id')
    ).order_by('created_day')
    
    complaints_data = defaultdict(int)
    for item in complaints_by_day:
        complaints_data[item['created_day'].strftime('%Y-%m-%d')] = item['count']
    
    # Prepare time series data for charts
    problems_time_series = {
        'dates': [d.strftime('%Y-%m-%d') for d in date_range],
        'series': [{
            'name': 'Problèmes',
            'data': [problems_data.get(d.strftime('%Y-%m-%d'), 0) for d in date_range]
        }]
    }
    
    complaints_time_series = {
        'dates': [d.strftime('%Y-%m-%d') for d in date_range],
        'series': [{
            'name': 'Réclamations',
            'data': [complaints_data.get(d.strftime('%Y-%m-%d'), 0) for d in date_range]
        }]
    }
    
    # Resolution time data
    # Calculate average time to resolve by category
    # First, get problems that have been resolved
    resolved_problems = Problem.objects.filter(
        municipality=municipality,
        status='RESOLVED'
    )
    
    # Group them by category and calculate average resolution time
    resolution_time_data = []
    
    # Get status logs for resolved problems
    if resolved_problems.exists():
        status_logs = StatusLog.objects.filter(
            record_type='PROBLEM',
            record_id__in=resolved_problems.values_list('id', flat=True),
            new_status='RESOLVED'
        )
        
        # Group by category
        category_resolution_times = defaultdict(list)
        
        for problem in resolved_problems:
            # Find the log entry where this problem was resolved
            log_entry = status_logs.filter(record_id=problem.id).first()
            
            if log_entry:
                # Calculate time difference in days
                resolution_time = (log_entry.changed_at.date() - problem.created_at.date()).days
                category_name = problem.category.name if problem.category else 'Sans catégorie'
                category_resolution_times[category_name].append(resolution_time)
        
        # Calculate averages
        for category, times in category_resolution_times.items():
            if times:
                avg_days = sum(times) / len(times)
                resolution_time_data.append({
                    'category': category,
                    'avg_days': avg_days
                })
    
    # Sort by average resolution time (ascending)
    resolution_time_data.sort(key=lambda x: x['avg_days'])
    
    # Get problem locations for the map
    problem_map_data = Problem.objects.filter(
        municipality=municipality
    ).select_related('category').values(
        'pk', 'latitude', 'longitude', 'description', 'status', 
        'created_at', 'category__name'
    )
    
    # Enhanced recent activities with more detail
    recent_activities = []
    
    # Recent problems with more detailed info
    recent_problems = Problem.objects.filter(
        municipality=municipality
    ).select_related('category', 'citizen').order_by('-created_at')[:10]
    
    for problem in recent_problems:
        status_display = dict(Problem.STATUS_CHOICES).get(problem.status, problem.status)
        detail_url = f"/admin/problems/{problem.id}/detail/"
        
        recent_activities.append({
            'type': 'problem',
            'text': f"Problème signalé: {problem.description[:50]}...",
            'timestamp': problem.created_at,
            'status': problem.status,
            'status_display': status_display,
            'detail_url': detail_url
        })
    
    # Recent complaints with more detailed info
    recent_complaints = Complaint.objects.filter(
        municipality=municipality
    ).select_related('citizen').order_by('-created_at')[:10]
    
    for complaint in recent_complaints:
        status_display = dict(Complaint.STATUS_CHOICES).get(complaint.status, complaint.status)
        detail_url = f"/admin/complaints/{complaint.id}/detail/"
        
        recent_activities.append({
            'type': 'complaint',
            'text': f"Réclamation soumise: {complaint.subject[:50]}...",
            'timestamp': complaint.created_at,
            'status': complaint.status,
            'status_display': status_display,
            'detail_url': detail_url
        })
    
    # Sort recent activities by timestamp
    recent_activities = sorted(recent_activities, key=lambda x: x['timestamp'], reverse=True)[:15]
    
    # Serialize data for JavaScript
    problem_status_data_dump = json.dumps(list(problem_status_data))
    complaint_status_data_dump = json.dumps(list(complaint_status_data))
    problem_categories_data_dump = json.dumps( list(problem_categories_data))
    problem_map_data_dump = json.dumps(list(problem_map_data), default=str)
    
    # Context for the template
    context = {
        'total_problems': total_problems,
        'new_problems_last_7_days': new_problems_last_period,
        'problem_percent_change': problem_percent_change,
        'new_problem_percent_change': new_problem_percent_change,
        
        'total_complaints': total_complaints,
        'new_complaints_last_7_days': new_complaints_last_period,
        'complaint_percent_change': complaint_percent_change,
        'new_complaint_percent_change': new_complaint_percent_change,
        
        'problem_status_data': problem_status_data_dump,
        'complaint_status_data': complaint_status_data_dump,
        'problem_categories_data': problem_categories_data_dump,
        'problems_time_series': json.dumps(problems_time_series),
        'complaints_time_series': json.dumps(complaints_time_series),
        'resolution_time_data': json.dumps(resolution_time_data),
        'problem_map_data': problem_map_data_dump,
        
        'recent_activities': recent_activities,
        'municipality': municipality,
        'navName': 'dashboard',
    }
    
    return render(request, 'muni_admin/dashboard.html', context)


from django.shortcuts import get_object_or_404
from Citoyen.models import Problem, StatusLog # Assuming models are in Citoyen app
from Citoyen.models import Notification
@login_required
def problem_detail(request, problem_id):
    # Fetch the specific problem, ensuring it belongs to the admin's municipality (or handle superadmin case)
    problem = get_object_or_404(Problem.objects.select_related(
        'citizen', 'category', 'municipality'
    ), pk=problem_id, municipality=request.user.admin_profile.municipality)
    
    # Fetch status change history for this problem
    status_logs = StatusLog.objects.filter(
        record_type='PROBLEM', 
        record_id=problem.id
    ).order_by('-changed_at').select_related('changed_by')
    
    # Handle status update form submission
    if request.method == 'POST':
        old_status = problem.status
        new_status = request.POST.get('status')
        comment = request.POST.get('comment', '')
        
        if new_status and new_status != old_status:
            # Update problem status
            problem.status = new_status
            problem.comment = comment
            problem.save()
            
            # Create status log entry
            StatusLog.objects.create(
                record_type='PROBLEM',
                record_id=problem.id,
                old_status=old_status,
                new_status=new_status,
                changed_by=request.user,
                changed_at=timezone.now()
            )
            
            # Optional: Create notification for the citizen
            Notification.objects.create(
                user=problem.citizen.user,
                title=_("Mise à jour de votre signalement"),
                message=_("Le statut de votre problème a été mis à jour à: {}").format(
                    dict(Problem.STATUS_CHOICES)[new_status]
                ),
                type='PROBLEM_UPDATE',
                related_id=str(problem.id),
                related_type='PROBLEM'
            )
            
            # Add success message
            from django.contrib import messages
            messages.success(request, _("Statut mis à jour avec succès."))
            
            # Redirect to avoid form resubmission
            from django.shortcuts import redirect
            return redirect('Muni_admin:problem_detail', problem_id=problem_id)
    
    # Prepare context for the template
    context = {
        'problem': problem,
        'status_logs': status_logs,
        'status_choices': Problem.STATUS_CHOICES,  # Pass status choices for potential status change form
        'navName': 'problems',  # For navigation highlighting
    }
    
    return render(request, 'muni_admin/problem_detail.html', context)



# Enhanced Report Generation Views - Complete Implementation
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.utils import timezone
import datetime
import json
import os
import tempfile
import subprocess
from collections import defaultdict
import base64
from io import BytesIO

# Chart and PDF libraries
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.patches import Wedge
import seaborn as sns
import folium
from folium.plugins import HeatMap, MarkerCluster
import plotly.graph_objects as go
import plotly.express as px
from plotly.offline import plot
import plotly.io as pio

# PDF libraries
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak, KeepTogether
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.graphics.shapes import Drawing, Rect
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics import renderPDF

# Import your models (adjust imports based on your project structure)
# from .models import Problem, Complaint, Category, StatusLog, Municipality

@login_required
def report_dashboard(request):
    """Dashboard for report generation with various options"""
    municipality = request.user.admin_profile.municipality
    
    # Get available date ranges
    earliest_problem = Problem.objects.filter(municipality=municipality).order_by('created_at').first()
    earliest_complaint = Complaint.objects.filter(municipality=municipality).order_by('created_at').first()
    
    earliest_date = None
    if earliest_problem and earliest_complaint:
        earliest_date = min(earliest_problem.created_at, earliest_complaint.created_at).date()
    elif earliest_problem:
        earliest_date = earliest_problem.created_at.date()
    elif earliest_complaint:
        earliest_date = earliest_complaint.created_at.date()
    
    context = {
        'municipality': municipality,
        'earliest_date': earliest_date,
        'categories': Category.objects.all(),
        'navName': 'reports',
    }
    
    return render(request, 'muni_admin/reports/dashboard.html', context)

@login_required
def generate_report(request):
    """Generate a comprehensive report based on selected parameters"""
    if request.method != 'POST':
        return redirect('Muni_admin:report_dashboard')
    
    municipality = request.user.admin_profile.municipality
    
    # Get report parameters
    report_type = request.POST.get('report_type', 'comprehensive')
    date_from = request.POST.get('date_from')
    date_to = request.POST.get('date_to')
    include_problems = request.POST.get('include_problems') == 'on'
    include_complaints = request.POST.get('include_complaints') == 'on'
    include_charts = request.POST.get('include_charts') == 'on'
    include_maps = request.POST.get('include_maps') == 'on'
    categories = request.POST.getlist('categories')
    format_type = request.POST.get('format', 'pdf')
    
    # Parse dates
    try:
        if date_from:
            date_from = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
        if date_to:
            date_to = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)
    
    # Generate report data
    report_data = _generate_report_data(
        municipality, date_from, date_to, include_problems, 
        include_complaints, categories
    )
    
    # Generate report based on format
    if format_type == 'pdf':
        return _generate_enhanced_pdf_report(request, report_data, include_charts, include_maps)
    elif format_type == 'excel':
        return _generate_excel_report(report_data)
    else:
        return JsonResponse({'error': 'Unsupported format'}, status=400)

def _generate_report_data(municipality, date_from, date_to, include_problems, include_complaints, categories):
    """Generate comprehensive report data with enhanced analytics"""
    data = {
        'municipality': municipality,
        'date_from': date_from,
        'date_to': date_to,
        'generated_at': timezone.now(),
        'generated_by': municipality.name,
    }
    
    # Base querysets
    problems_qs = Problem.objects.filter(municipality=municipality)
    complaints_qs = Complaint.objects.filter(municipality=municipality)
    
    # Apply date filters
    if date_from:
        problems_qs = problems_qs.filter(created_at__date__gte=date_from)
        complaints_qs = complaints_qs.filter(created_at__date__gte=date_from)
    
    if date_to:
        problems_qs = problems_qs.filter(created_at__date__lte=date_to)
        complaints_qs = complaints_qs.filter(created_at__date__lte=date_to)
    
    # Apply category filters
    if categories:
        problems_qs = problems_qs.filter(category_id__in=categories)
    
    # Enhanced Problems data
    if include_problems:
        problems_by_status = list(problems_qs.values('status').annotate(count=Count('status')))
        problems_by_category = list(problems_qs.values('category__name').annotate(count=Count('category')))
        
        data['problems'] = {
            'total': problems_qs.count(),
            'by_status': problems_by_status,
            'by_category': problems_by_category,
            'recent': list(problems_qs.order_by('-created_at')[:20].values(
                'id', 'description', 'location', 'status', 'created_at', 'category__name', 'latitude', 'longitude'
            )),
            'resolution_stats': _calculate_resolution_stats(problems_qs),
            'priority_analysis': _analyze_problem_priorities(problems_qs),
            'geographic_distribution': _get_geographic_distribution(problems_qs),
        }
    
    # Enhanced Complaints data
    if include_complaints:
        complaints_by_status = list(complaints_qs.values('status').annotate(count=Count('status')))
        
        data['complaints'] = {
            'total': complaints_qs.count(),
            'by_status': complaints_by_status,
            'recent': list(complaints_qs.order_by('-created_at')[:20].values(
                'id', 'subject', 'description', 'status', 'created_at'
            )),
            'response_time_analysis': _analyze_complaint_response_times(complaints_qs),
        }
    
    # Enhanced Time series data
    data['time_series'] = _generate_enhanced_time_series_data(problems_qs, complaints_qs, date_from, date_to)
    
    # Enhanced Summary statistics
    total_problems = problems_qs.count() if include_problems else 0
    total_complaints = complaints_qs.count() if include_complaints else 0
    resolved_problems = problems_qs.filter(status='RESOLVED').count() if include_problems else 0
    resolved_complaints = complaints_qs.filter(status='RESOLVED').count() if include_complaints else 0
    pending_problems = problems_qs.filter(status='PENDING').count() if include_problems else 0
    pending_complaints = complaints_qs.filter(status='PENDING').count() if include_complaints else 0
    
    data['summary'] = {
        'total_issues': total_problems + total_complaints,
        'total_problems': total_problems,
        'total_complaints': total_complaints,
        'resolved_issues': resolved_problems + resolved_complaints,
        'resolved_problems': resolved_problems,
        'resolved_complaints': resolved_complaints,
        'pending_issues': pending_problems + pending_complaints,
        'pending_problems': pending_problems,
        'pending_complaints': pending_complaints,
        'in_progress_problems': problems_qs.filter(status='IN_PROGRESS').count() if include_problems else 0,
        'reviewing_complaints': complaints_qs.filter(status='REVIEWING').count() if include_complaints else 0,
    }
    
    if data['summary']['total_issues'] > 0:
        data['summary']['resolution_rate'] = round(
            (data['summary']['resolved_issues'] / data['summary']['total_issues']) * 100, 2
        )
        data['summary']['pending_rate'] = round(
            (data['summary']['pending_issues'] / data['summary']['total_issues']) * 100, 2
        )
    else:
        data['summary']['resolution_rate'] = 0
        data['summary']['pending_rate'] = 0
    
    # Performance metrics
    data['performance'] = _calculate_performance_metrics(problems_qs, complaints_qs, date_from, date_to)
    
    return data

def _analyze_problem_priorities(problems_qs):
    """Analyze problem priorities based on various factors"""
    high_priority = problems_qs.filter(
        Q(status='PENDING') & Q(created_at__gte=timezone.now() - datetime.timedelta(days=7))
    ).count()
    
    medium_priority = problems_qs.filter(
        Q(status='IN_PROGRESS') | (Q(status='PENDING') & Q(created_at__gte=timezone.now() - datetime.timedelta(days=30)))
    ).count()
    
    low_priority = problems_qs.exclude(
        Q(status='PENDING') | Q(status='IN_PROGRESS')
    ).count()
    
    return {
        'high': high_priority,
        'medium': medium_priority,
        'low': low_priority
    }

def _get_geographic_distribution(problems_qs):
    """Get geographic distribution of problems"""
    locations = []
    for problem in problems_qs.filter(latitude__isnull=False, longitude__isnull=False):
        locations.append({
            'lat': problem.latitude,
            'lng': problem.longitude,
            'status': problem.status,
            'category': problem.category.name if problem.category else 'Sans catégorie',
            'description': problem.description[:100] + '...' if len(problem.description) > 100 else problem.description
        })
    return locations

def _analyze_complaint_response_times(complaints_qs):
    """Analyze complaint response times"""
    response_times = []
    
    for complaint in complaints_qs.filter(status__in=['RESOLVED', 'REJECTED']):
        log_entry = StatusLog.objects.filter(
            record_type='COMPLAINT',
            record_id=complaint.id,
            new_status__in=['RESOLVED', 'REJECTED']
        ).first()
        
        if log_entry:
            response_time = (log_entry.changed_at.date() - complaint.created_at.date()).days
            response_times.append(response_time)
    
    if response_times:
        avg_response_time = sum(response_times) / len(response_times)
        min_response_time = min(response_times)
        max_response_time = max(response_times)
    else:
        avg_response_time = min_response_time = max_response_time = 0
    
    return {
        'average': round(avg_response_time, 1),
        'minimum': min_response_time,
        'maximum': max_response_time,
        'count': len(response_times)
    }

def _calculate_performance_metrics(problems_qs, complaints_qs, date_from, date_to):
    """Calculate performance metrics for the municipality"""
    current_period_days = (date_to - date_from).days if date_from and date_to else 30
    
    # Calculate previous period for comparison
    if date_from and date_to:
        prev_date_to = date_from - datetime.timedelta(days=1)
        prev_date_from = prev_date_to - datetime.timedelta(days=current_period_days)
        
        prev_problems = Problem.objects.filter(
            municipality=problems_qs.first().municipality if problems_qs.exists() else None,
            created_at__date__gte=prev_date_from,
            created_at__date__lte=prev_date_to
        ).count()
        
        prev_complaints = Complaint.objects.filter(
            municipality=complaints_qs.first().municipality if complaints_qs.exists() else None,
            created_at__date__gte=prev_date_from,
            created_at__date__lte=prev_date_to
        ).count()
    else:
        prev_problems = prev_complaints = 0
    
    current_problems = problems_qs.count()
    current_complaints = complaints_qs.count()
    
    # Calculate trends
    problems_trend = ((current_problems - prev_problems) / prev_problems * 100) if prev_problems > 0 else 0
    complaints_trend = ((current_complaints - prev_complaints) / prev_complaints * 100) if prev_complaints > 0 else 0
    
    return {
        'current_problems': current_problems,
        'current_complaints': current_complaints,
        'previous_problems': prev_problems,
        'previous_complaints': prev_complaints,
        'problems_trend': round(problems_trend, 1),
        'complaints_trend': round(complaints_trend, 1),
        'daily_average_problems': round(current_problems / current_period_days, 1) if current_period_days > 0 else 0,
        'daily_average_complaints': round(current_complaints / current_period_days, 1) if current_period_days > 0 else 0,
    }

def _calculate_resolution_stats(problems_qs):
    """Calculate enhanced resolution statistics for problems"""
    resolved_problems = problems_qs.filter(status='RESOLVED')
    
    if not resolved_problems.exists():
        return {
            'avg_resolution_time': 0, 
            'by_category': [],
            'resolution_trend': [],
            'fastest_resolution': 0,
            'slowest_resolution': 0
        }
    
    # Get resolution times from status logs
    resolution_times = []
    category_times = defaultdict(list)
    monthly_resolutions = defaultdict(int)
    
    for problem in resolved_problems:
        log_entry = StatusLog.objects.filter(
            record_type='PROBLEM',
            record_id=problem.id,
            new_status='RESOLVED'
        ).first()
        
        if log_entry:
            resolution_time = (log_entry.changed_at.date() - problem.created_at.date()).days
            resolution_times.append(resolution_time)
            
            category_name = problem.category.name if problem.category else 'Sans catégorie'
            category_times[category_name].append(resolution_time)
            
            # Monthly resolution tracking
            month_key = log_entry.changed_at.strftime('%Y-%m')
            monthly_resolutions[month_key] += 1
    
    avg_resolution_time = sum(resolution_times) / len(resolution_times) if resolution_times else 0
    
    by_category = []
    for category, times in category_times.items():
        by_category.append({
            'category': category,
            'avg_days': round(sum(times) / len(times), 1),
            'count': len(times),
            'min_days': min(times),
            'max_days': max(times)
        })
    
    # Resolution trend over months
    resolution_trend = [{'month': month, 'count': count} for month, count in sorted(monthly_resolutions.items())]
    
    return {
        'avg_resolution_time': round(avg_resolution_time, 1),
        'by_category': sorted(by_category, key=lambda x: x['avg_days']),
        'resolution_trend': resolution_trend,
        'fastest_resolution': min(resolution_times) if resolution_times else 0,
        'slowest_resolution': max(resolution_times) if resolution_times else 0,
        'total_resolved': len(resolution_times)
    }

def _generate_enhanced_time_series_data(problems_qs, complaints_qs, date_from, date_to):
    """Generate enhanced time series data for charts"""
    if not date_from:
        date_from = timezone.now().date() - datetime.timedelta(days=30)
    if not date_to:
        date_to = timezone.now().date()
    
    # Generate date range
    date_range = []
    current_date = date_from
    while current_date <= date_to:
        date_range.append(current_date)
        current_date += datetime.timedelta(days=1)
    
    from django.db.models.functions import TruncDate
    
    # Problems by day with status breakdown
    problems_by_day = problems_qs.annotate(
        day=TruncDate('created_at')
    ).values('day', 'status').annotate(count=Count('id'))
    
    # Organize problems data by status
    problems_data = defaultdict(lambda: defaultdict(int))
    for item in problems_by_day:
        day_str = item['day'].strftime('%Y-%m-%d')
        problems_data[item['status']][day_str] = item['count']
    
    # Complaints by day
    complaints_by_day = complaints_qs.annotate(
        day=TruncDate('created_at')
    ).values('day').annotate(count=Count('id'))
    
    complaints_data = {item['day'].strftime('%Y-%m-%d'): item['count'] for item in complaints_by_day}
    
    # Build series data
    date_strings = [d.strftime('%Y-%m-%d') for d in date_range]
    
    return {
        'dates': date_strings,
        'problems': {
            'total': [sum(problems_data[status].get(d, 0) for status in problems_data.keys()) for d in date_strings],
            'pending': [problems_data['PENDING'].get(d, 0) for d in date_strings],
            'in_progress': [problems_data['IN_PROGRESS'].get(d, 0) for d in date_strings],
            'resolved': [problems_data['RESOLVED'].get(d, 0) for d in date_strings],
            'rejected': [problems_data['REJECTED'].get(d, 0) for d in date_strings],
        },
        'complaints': [complaints_data.get(d, 0) for d in date_strings]
    }

# Enhanced Chart Generation Functions
def _generate_chart_image(chart_type, data, title, filename, width=800, height=600):
    """Generate chart images using matplotlib with enhanced styling"""
    plt.style.use('seaborn-v0_8')
    fig, ax = plt.subplots(figsize=(width/100, height/100), dpi=100)
    
    # Set color palette
    colors_palette = ['#3498db', '#e74c3c', '#2ecc71', '#f39c12', '#9b59b6', '#1abc9c', '#34495e', '#e67e22']
    
    if chart_type == 'pie':
        # Pie chart for status distribution
        labels = [item['status'] for item in data]
        sizes = [item['count'] for item in data]
        
        # Translate status labels to French
        status_translation = {
            'PENDING': 'En attente',
            'IN_PROGRESS': 'En cours',
            'RESOLVED': 'Résolu',
            'REJECTED': 'Rejeté',
            'REVIEWING': 'En examen',
            'DELEGATED': 'Délégué'
        }
        labels = [status_translation.get(label, label) for label in labels]
        
        wedges, texts, autotexts = ax.pie(sizes, labels=labels, autopct='%1.1f%%', 
                                         colors=colors_palette[:len(labels)],
                                         startangle=90, explode=[0.05]*len(labels))
        
        # Enhance text styling
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontweight('bold')
            autotext.set_fontsize(10)
        
        for text in texts:
            text.set_fontsize(11)
            text.set_fontweight('bold')
    
    elif chart_type == 'bar':
        # Bar chart for categories
        categories = [item['category__name'] or 'Sans catégorie' for item in data]
        counts = [item['count'] for item in data]
        
        bars = ax.bar(categories, counts, color=colors_palette[:len(categories)])
        
        # Add value labels on bars
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                   f'{int(height)}', ha='center', va='bottom', fontweight='bold')
        
        ax.set_xlabel('Catégories', fontweight='bold', fontsize=12)
        ax.set_ylabel('Nombre de signalements', fontweight='bold', fontsize=12)
        plt.xticks(rotation=45, ha='right')
    
    elif chart_type == 'line':
        # Line chart for time series
        dates = data['dates']
        problems_total = data['problems']['total']
        complaints = data['complaints']
        
        # Convert date strings to datetime objects
        date_objects = [datetime.datetime.strptime(d, '%Y-%m-%d').date() for d in dates]
        
        ax.plot(date_objects, problems_total, marker='o', linewidth=2, 
               label='Problèmes', color='#e74c3c', markersize=4)
        ax.plot(date_objects, complaints, marker='s', linewidth=2, 
               label='Réclamations', color='#3498db', markersize=4)
        
        ax.set_xlabel('Date', fontweight='bold', fontsize=12)
        ax.set_ylabel('Nombre de signalements', fontweight='bold', fontsize=12)
        ax.legend(fontsize=11)
        ax.grid(True, alpha=0.3)
        
        # Format x-axis dates
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
        ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates)//10)))
        plt.xticks(rotation=45)
    
    elif chart_type == 'stacked_bar':
        # Stacked bar chart for problem status over time
        dates = data['dates']
        pending = data['problems']['pending']
        in_progress = data['problems']['in_progress']
        resolved = data['problems']['resolved']
        rejected = data['problems']['rejected']
        
        # Convert dates for plotting
        date_objects = [datetime.datetime.strptime(d, '%Y-%m-%d').date() for d in dates]
        
        ax.bar(date_objects, pending, label='En attente', color='#f39c12')
        ax.bar(date_objects, in_progress, bottom=pending, label='En cours', color='#3498db')
        
        bottom_resolved = [p + ip for p, ip in zip(pending, in_progress)]
        ax.bar(date_objects, resolved, bottom=bottom_resolved, label='Résolu', color='#2ecc71')
        
        bottom_rejected = [br + r for br, r in zip(bottom_resolved, resolved)]
        ax.bar(date_objects, rejected, bottom=bottom_rejected, label='Rejeté', color='#e74c3c')
        
        ax.set_xlabel('Date', fontweight='bold', fontsize=12)
        ax.set_ylabel('Nombre de problèmes', fontweight='bold', fontsize=12)
        ax.legend(fontsize=10)
        ax.grid(True, alpha=0.3, axis='y')
        
        # Format x-axis
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
        ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates)//10)))
        plt.xticks(rotation=45)
    
    # Set title and styling
    ax.set_title(title, fontsize=16, fontweight='bold', pad=20)
    
    # Improve layout
    plt.tight_layout()
    
    # Save the chart
    plt.savefig(filename, dpi=150, bbox_inches='tight', facecolor='white', edgecolor='none')
    plt.close()
    
    return filename

def _generate_map_image(locations, municipality_name, filename, width=800, height=600):
    """Generate map visualization using folium"""
    if not locations:
        # Create empty map centered on a default location
        center_lat, center_lng = 18.0735, -15.9582  # Mauritania center
        m = folium.Map(location=[center_lat, center_lng], zoom_start=6)
        
        # Add a message for no data
        folium.Marker(
            [center_lat, center_lng],
            popup="Aucune donnée géographique disponible",
            icon=folium.Icon(color='gray', icon='info-sign')
        ).add_to(m)
    else:
        # Calculate center point
        center_lat = sum(loc['lat'] for loc in locations) / len(locations)
        center_lng = sum(loc['lng'] for loc in locations) / len(locations)
        
        # Create map
        m = folium.Map(location=[center_lat, center_lng], zoom_start=12)
        
        # Color mapping for status
        status_colors = {
            'PENDING': 'orange',
            'IN_PROGRESS': 'blue',
            'RESOLVED': 'green',
            'REJECTED': 'red',
            'DELEGATED': 'purple'
        }
        
        # Add markers with clustering
        marker_cluster = MarkerCluster().add_to(m)
        
        for loc in locations:
            color = status_colors.get(loc['status'], 'gray')
            
            folium.Marker(
                [loc['lat'], loc['lng']],
                popup=f"""
                <div style="width: 200px;">
                    <b>Catégorie:</b> {loc['category']}<br>
                    <b>Statut:</b> {loc['status']}<br>
                    <b>Description:</b> {loc['description']}
                </div>
                """,
                icon=folium.Icon(color=color, icon='exclamation-sign')
            ).add_to(marker_cluster)
        
        # Add heat map layer
        heat_data = [[loc['lat'], loc['lng']] for loc in locations]
        HeatMap(heat_data, radius=15, blur=10, max_zoom=1).add_to(m)
    
    # Add title
    title_html = f'''
    <h3 align="center" style="font-size:16px; margin-top:0;">
        <b>Distribution Géographique des Problèmes - {municipality_name}</b>
    </h3>
    '''
    m.get_root().html.add_child(folium.Element(title_html))
    
    # Save map as HTML first, then convert to image
    temp_html = filename.replace('.png', '.html')
    m.save(temp_html)
    
    # For now, return the HTML file path since converting to image requires additional setup
    # In production, you might want to use selenium or similar to convert HTML to image
    return temp_html

def _generate_enhanced_pdf_report(request, report_data, include_charts, include_maps):
    """Generate enhanced PDF report with charts and maps"""
    try:
        # Create temporary directory for charts and maps
        temp_dir = tempfile.mkdtemp()
        chart_files = []
        map_files = []
        
        # Generate charts if requested
        if include_charts:
            # 1. Problems by Status (Pie Chart)
            if 'problems' in report_data and report_data['problems']['by_status']:
                chart_file = os.path.join(temp_dir, 'problems_status_pie.png')
                _generate_chart_image(
                    'pie', 
                    report_data['problems']['by_status'],
                    'Répartition des Problèmes par Statut',
                    chart_file
                )
                chart_files.append(('problems_status', chart_file))
            
            # 2. Problems by Category (Bar Chart)
            if 'problems' in report_data and report_data['problems']['by_category']:
                chart_file = os.path.join(temp_dir, 'problems_category_bar.png')
                _generate_chart_image(
                    'bar',
                    report_data['problems']['by_category'],
                    'Problèmes par Catégorie',
                    chart_file
                )
                chart_files.append(('problems_category', chart_file))
            
            # 3. Time Series (Line Chart)
            if report_data['time_series']['dates']:
                chart_file = os.path.join(temp_dir, 'time_series_line.png')
                _generate_chart_image(
                    'line',
                    report_data['time_series'],
                    'Évolution Temporelle des Signalements',
                    chart_file
                )
                chart_files.append(('time_series', chart_file))
            
            # 4. Stacked Bar Chart for Problem Status Over Time
            if report_data['time_series']['dates']:
                chart_file = os.path.join(temp_dir, 'problems_status_stacked.png')
                _generate_chart_image(
                    'stacked_bar',
                    report_data['time_series'],
                    'Évolution des Statuts des Problèmes',
                    chart_file
                )
                chart_files.append(('problems_status_time', chart_file))
            
            # 5. Complaints by Status (Pie Chart)
            if 'complaints' in report_data and report_data['complaints']['by_status']:
                chart_file = os.path.join(temp_dir, 'complaints_status_pie.png')
                _generate_chart_image(
                    'pie',
                    report_data['complaints']['by_status'],
                    'Répartition des Réclamations par Statut',
                    chart_file
                )
                chart_files.append(('complaints_status', chart_file))
        
        # Generate maps if requested
        if include_maps and 'problems' in report_data:
            map_file = os.path.join(temp_dir, 'geographic_distribution.png')
            map_html = _generate_map_image(
                report_data['problems'].get('geographic_distribution', []),
                report_data['municipality'].name,
                map_file
            )
            map_files.append(('geographic_map', map_html))
        
        # Create PDF using ReportLab
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            topMargin=1*inch, 
            bottomMargin=1*inch,
            leftMargin=0.75*inch,
            rightMargin=0.75*inch
        )
        
        # Enhanced styles
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#34495e'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2c3e50'),
            spaceBefore=25,
            spaceAfter=15,
            fontName='Helvetica-Bold'
        )
        
        subheading_style = ParagraphStyle(
            'CustomSubheading',
            parent=styles['Heading3'],
            fontSize=14,
            textColor=colors.HexColor('#34495e'),
            spaceBefore=20,
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=8,
            fontName='Helvetica'
        )
        
        # Build story (content)
        story = []
        
        # Cover Page
        story.append(Spacer(1, 2*inch))
        story.append(Paragraph("RAPPORT MUNICIPAL", title_style))
        logo_path = None
        # Try to find the logo in STATIC_ROOT (for production)
        if settings.STATIC_ROOT:
            potential_path = os.path.join(settings.STATIC_ROOT, 'municipal_logo.jpg')
            if os.path.exists(potential_path):
                logo_path = potential_path
        from django.contrib.staticfiles.finders import find # Import find for development environments
        # If not found in STATIC_ROOT, try to find it using Django's staticfiles finders (for development)
        if not logo_path:
            logo_path = find('municipal_logo.jpg')

        if logo_path:
            try:
                # Adjust width and height as needed
                logo = Image(logo_path, width=2*inch, height=2*inch) 
                story.append(logo)
                story.append(Spacer(1, 0.5*inch)) # Add some space after the logo
            except Exception as e:
                # Log the error if the image cannot be loaded
                print(f"Error loading logo image: {e}")
                story.append(Paragraph("Logo non disponible", normal_style)) # Placeholder text
                story.append(Spacer(1, 0.5*inch))
        else:
            story.append(Paragraph("Logo non trouvé (chemin: municipal_logo.jpg)", normal_style))
            story.append(Spacer(1, 0.5*inch))
        
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph(f"Municipalité de {report_data['municipality'].name}", subtitle_style))
        story.append(Spacer(1, 1*inch))
        
        # Report period
        if report_data['date_from'] and report_data['date_to']:
            period_text = f"Période: {report_data['date_from'].strftime('%d/%m/%Y')} - {report_data['date_to'].strftime('%d/%m/%Y')}"
        else:
            period_text = "Période: Toutes les données disponibles"
        story.append(Paragraph(period_text, subtitle_style))
        
        story.append(Spacer(1, 1*inch))
        story.append(Paragraph(f"Généré le {report_data['generated_at'].strftime('%d/%m/%Y à %H:%M')}", normal_style))
        story.append(PageBreak())
        
        # Executive Summary
        story.append(Paragraph("RÉSUMÉ EXÉCUTIF", heading_style))
        
        summary_data = [
            ['Indicateur', 'Valeur'],
            ['Total des signalements', str(report_data['summary']['total_issues'])],
            ['Problèmes signalés', str(report_data['summary']['total_problems'])],
            ['Réclamations déposées', str(report_data['summary']['total_complaints'])],
            ['Signalements résolus', str(report_data['summary']['resolved_issues'])],
            ['Taux de résolution', f"{report_data['summary']['resolution_rate']}%"],
            ['Signalements en attente', str(report_data['summary']['pending_issues'])],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ecf0f1')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Performance Metrics
        if 'performance' in report_data:
            story.append(Paragraph("INDICATEURS DE PERFORMANCE", subheading_style))
            
            perf = report_data['performance']
            performance_text = f"""
            <b>Tendances:</b><br/>
            • Problèmes: {perf['problems_trend']:+.1f}% par rapport à la période précédente<br/>
            • Réclamations: {perf['complaints_trend']:+.1f}% par rapport à la période précédente<br/>
            <br/>
            <b>Moyennes quotidiennes:</b><br/>
            • Problèmes: {perf['daily_average_problems']} par jour<br/>
            • Réclamations: {perf['daily_average_complaints']} par jour
            """
            story.append(Paragraph(performance_text, normal_style))
            story.append(Spacer(1, 20))
        
        # Charts Section
        if include_charts and chart_files:
            story.append(PageBreak())
            story.append(Paragraph("ANALYSES GRAPHIQUES", heading_style))
            
            for chart_name, chart_file in chart_files:
                if os.path.exists(chart_file):
                    try:
                        # Add chart image
                        img = Image(chart_file, width=6*inch, height=4.5*inch)
                        story.append(img)
                        story.append(Spacer(1, 20))
                    except Exception as e:
                        story.append(Paragraph(f"Erreur lors du chargement du graphique: {chart_name}", normal_style))
                        story.append(Spacer(1, 10))
        
        # Problems Section
        if 'problems' in report_data:
            story.append(PageBreak())
            story.append(Paragraph("ANALYSE DES PROBLÈMES", heading_style))
            
            problems = report_data['problems']
            
            # Problems summary
            story.append(Paragraph("Résumé des Problèmes", subheading_style))
            problems_summary = f"""
            <b>Total des problèmes:</b> {problems['total']}<br/>
            <b>Temps moyen de résolution:</b> {problems['resolution_stats']['avg_resolution_time']} jours<br/>
            <b>Résolution la plus rapide:</b> {problems['resolution_stats']['fastest_resolution']} jours<br/>
            <b>Résolution la plus lente:</b> {problems['resolution_stats']['slowest_resolution']} jours
            """
            story.append(Paragraph(problems_summary, normal_style))
            story.append(Spacer(1, 15))
            
            # Priority Analysis
            if 'priority_analysis' in problems:
                story.append(Paragraph("Analyse des Priorités", subheading_style))
                priority = problems['priority_analysis']
                priority_data = [
                    ['Priorité', 'Nombre'],
                    ['Haute (< 7 jours)', str(priority['high'])],
                    ['Moyenne (7-30 jours)', str(priority['medium'])],
                    ['Basse (> 30 jours)', str(priority['low'])],
                ]
                
                priority_table = Table(priority_data, colWidths=[2.5*inch, 1.5*inch])
                priority_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fadbd8')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#c0392b')),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                ]))
                
                story.append(priority_table)
                story.append(Spacer(1, 15))
            
            # Recent Problems Table
            if problems['recent']:
                story.append(Paragraph("Problèmes Récents", subheading_style))
                
                recent_data = [['ID', 'Description', 'Localisation', 'Statut', 'Date']]
                for problem in problems['recent'][:10]:  # Limit to 10 most recent
                    recent_data.append([
                        str(problem['id'])[:8] + '...',
                        (problem['description'][:40] + '...') if len(problem['description']) > 40 else problem['description'],
                        problem['location'] or 'N/A',
                        problem['status'],
                        problem['created_at'].strftime('%d/%m/%Y')
                    ])
                
                recent_table = Table(recent_data, colWidths=[1*inch, 2.5*inch, 1.5*inch, 1*inch, 1*inch])
                recent_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#d5f4e6')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#27ae60')),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                
                story.append(recent_table)
                story.append(Spacer(1, 15))
        
        # Complaints Section
        if 'complaints' in report_data:
            story.append(PageBreak())
            story.append(Paragraph("ANALYSE DES RÉCLAMATIONS", heading_style))
            
            complaints = report_data['complaints']
            
            # Complaints summary
            story.append(Paragraph("Résumé des Réclamations", subheading_style))
            complaints_summary = f"""
            <b>Total des réclamations:</b> {complaints['total']}<br/>
            <b>Temps moyen de réponse:</b> {complaints['response_time_analysis']['average']} jours<br/>
            <b>Réponse la plus rapide:</b> {complaints['response_time_analysis']['minimum']} jours<br/>
            <b>Réponse la plus lente:</b> {complaints['response_time_analysis']['maximum']} jours
            """
            story.append(Paragraph(complaints_summary, normal_style))
            story.append(Spacer(1, 15))
            
            # Recent Complaints Table
            if complaints['recent']:
                story.append(Paragraph("Réclamations Récentes", subheading_style))
                
                recent_complaints_data = [['ID', 'Sujet', 'Description', 'Statut', 'Date']]
                for complaint in complaints['recent'][:10]:
                    recent_complaints_data.append([
                        str(complaint['id'])[:8] + '...',
                        (complaint['subject'][:30] + '...') if len(complaint['subject']) > 30 else complaint['subject'],
                        (complaint['description'][:40] + '...') if len(complaint['description']) > 40 else complaint['description'],
                        complaint['status'],
                        complaint['created_at'].strftime('%d/%m/%Y')
                    ])
                
                complaints_table = Table(recent_complaints_data, colWidths=[1*inch, 2*inch, 2.5*inch, 1*inch, 1*inch])
                complaints_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f39c12')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fdeaa7')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e67e22')),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                
                story.append(complaints_table)
                story.append(Spacer(1, 15))
        
        # Maps Section
        if include_maps and map_files:
            story.append(PageBreak())
            story.append(Paragraph("DISTRIBUTION GÉOGRAPHIQUE", heading_style))
            story.append(Paragraph("Carte de distribution des problèmes signalés dans la municipalité.", normal_style))
            story.append(Spacer(1, 10))
            
            for map_name, map_file in map_files:
                if os.path.exists(map_file):
                    story.append(Paragraph(f"Fichier de carte généré: {os.path.basename(map_file)}", normal_style))
                    story.append(Paragraph("Note: La carte interactive est disponible en format HTML séparé.", normal_style))
                    story.append(Spacer(1, 10))
        
        # Footer
        story.append(PageBreak())
        story.append(Spacer(1, 2*inch))
        story.append(Paragraph("CONCLUSION", heading_style))
        
        conclusion_text = f"""
        Ce rapport présente une analyse complète des signalements reçus par la municipalité de {report_data['municipality'].name}.
        Les données analysées couvrent {report_data['summary']['total_issues']} signalements au total, 
        avec un taux de résolution de {report_data['summary']['resolution_rate']}%.
        
        Les graphiques et analyses présentés dans ce rapport permettent d'identifier les tendances,
        les zones d'amélioration et les succès de la gestion municipale.
        """
        logo_path = None
        # Try to find the logo in STATIC_ROOT (for production)
        if settings.STATIC_ROOT:
            potential_path = os.path.join(settings.STATIC_ROOT, 'municipal_logo.jpg')
            if os.path.exists(potential_path):
                logo_path = potential_path
        from django.contrib.staticfiles.finders import find # Import find for development environments
        # If not found in STATIC_ROOT, try to find it using Django's staticfiles finders (for development)
        if not logo_path:
            logo_path = find('municipal_logo.jpg')

        if logo_path:
            try:
                # Adjust width and height as needed
                logo = Image(logo_path, width=2*inch, height=2*inch) 
                story.append(logo)
                story.append(Spacer(1, 0.5*inch)) # Add some space after the logo
            except Exception as e:
                # Log the error if the image cannot be loaded
                print(f"Error loading logo image: {e}")
                story.append(Paragraph("Logo non disponible", normal_style)) # Placeholder text
                story.append(Spacer(1, 0.5*inch))
        else:
            story.append(Paragraph("Logo non trouvé (chemin: municipal_logo.jpg)", normal_style))
            story.append(Spacer(1, 0.5*inch))
        
        story.append(Paragraph(conclusion_text, normal_style))
        story.append(Spacer(1, 30))
        story.append(Paragraph("Rapport généré automatiquement par le système de gestion municipale Belediyti", 
                              ParagraphStyle('Footer', parent=styles['Italic'], fontSize=9, textColor=colors.HexColor('#7f8c8d'))))
        # Build PDF
        doc.build(story)
        
        # Get PDF content
        pdf_content = buffer.getvalue()
        buffer.close()
        
        # Clean up temporary files
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        # Create response
        response = HttpResponse(pdf_content, content_type='application/pdf')
        filename = f"rapport_{report_data['municipality'].name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return JsonResponse({
            'error': f'PDF generation failed: {str(e)}',
            'details': error_details
        }, status=500)

def _generate_excel_report(report_data):
    """Generate Excel report with enhanced formatting"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, PieChart, LineChart, Reference
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        return JsonResponse({'error': 'Excel generation not available'}, status=500)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    subheader_font = Font(size=12, bold=True, color='2C3E50')
    subheader_fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
    normal_font = Font(size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    def auto_adjust_columns(worksheet):
        """Helper function to auto-adjust column widths while handling merged cells"""
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = None
            
            for cell in column_cells:
                # Skip merged cells
                if isinstance(cell, MergedCell):
                    continue
                    
                # Get column letter from first non-merged cell
                if column_letter is None:
                    column_letter = cell.column_letter
                
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width if we found a valid column letter
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Summary sheet
    ws_summary = wb.create_sheet("Résumé Exécutif")
    
    # Title
    ws_summary['A1'] = "RAPPORT MUNICIPAL"
    ws_summary['A1'].font = Font(size=18, bold=True, color='2C3E50')
    ws_summary.merge_cells('A1:D1')
    
    # Municipality info
    ws_summary['A3'] = "Municipalité:"
    ws_summary['B3'] = report_data['municipality'].name
    ws_summary['A4'] = "Période:"
    if report_data['date_from'] and report_data['date_to']:
        period_text = f"Du {report_data['date_from']} au {report_data['date_to']}"
    else:
        period_text = "Toutes les données"
    ws_summary['B4'] = period_text
    ws_summary['A5'] = "Généré le:"
    ws_summary['B5'] = report_data['generated_at'].strftime('%d/%m/%Y %H:%M')
    
    # Summary statistics
    row = 7
    ws_summary[f'A{row}'] = "STATISTIQUES GÉNÉRALES"
    ws_summary[f'A{row}'].font = subheader_font
    ws_summary[f'A{row}'].fill = subheader_fill
    ws_summary.merge_cells(f'A{row}:D{row}')
    
    row += 1
    summary_data = [
        ['Indicateur', 'Valeur', 'Pourcentage', 'Commentaire'],
        ['Total des signalements', report_data['summary']['total_issues'], '100%', 'Base de calcul'],
        ['Problèmes signalés', report_data['summary']['total_problems'], 
         f"{(report_data['summary']['total_problems']/max(report_data['summary']['total_issues'], 1)*100):.1f}%", 'Signalements de problèmes'],
        ['Réclamations déposées', report_data['summary']['total_complaints'],
         f"{(report_data['summary']['total_complaints']/max(report_data['summary']['total_issues'], 1)*100):.1f}%", 'Réclamations formelles'],
        ['Signalements résolus', report_data['summary']['resolved_issues'],
         f"{report_data['summary']['resolution_rate']}%", 'Taux de résolution global'],
        ['Signalements en attente', report_data['summary']['pending_issues'],
         f"{report_data['summary']['pending_rate']}%", 'En cours de traitement'],
    ]
    
    for i, row_data in enumerate(summary_data):
        for j, value in enumerate(row_data):
            cell = ws_summary.cell(row=row+i, column=j+1, value=value)
            if i == 0:  # Header row
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.font = normal_font
            cell.border = border
    
    # Auto-adjust column widths for summary sheet
    auto_adjust_columns(ws_summary)
    
    # Problems sheet
    if 'problems' in report_data:
        ws_problems = wb.create_sheet("Analyse des Problèmes")
        
        # Title
        ws_problems['A1'] = "ANALYSE DÉTAILLÉE DES PROBLÈMES"
        ws_problems['A1'].font = subheader_font
        ws_problems['A1'].fill = subheader_fill
        ws_problems.merge_cells('A1:F1')
        
        # Status distribution
        row = 3
        ws_problems[f'A{row}'] = "Répartition par Statut"
        ws_problems[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        status_headers = ['Statut', 'Nombre', 'Pourcentage']
        for j, header in enumerate(status_headers):
            cell = ws_problems.cell(row=row, column=j+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        total_problems = report_data['problems']['total']
        for item in report_data['problems']['by_status']:
            row += 1
            percentage = (item['count'] / max(total_problems, 1)) * 100
            status_data = [item['status'], item['count'], f"{percentage:.1f}%"]
            for j, value in enumerate(status_data):
                cell = ws_problems.cell(row=row, column=j+1, value=value)
                cell.font = normal_font
                cell.border = border
        
        # Category distribution
        row += 3
        ws_problems[f'A{row}'] = "Répartition par Catégorie"
        ws_problems[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        category_headers = ['Catégorie', 'Nombre', 'Pourcentage']
        for j, header in enumerate(category_headers):
            cell = ws_problems.cell(row=row, column=j+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for item in report_data['problems']['by_category']:
            row += 1
            percentage = (item['count'] / max(total_problems, 1)) * 100
            category_data = [item['category__name'] or 'Sans catégorie', item['count'], f"{percentage:.1f}%"]
            for j, value in enumerate(category_data):
                cell = ws_problems.cell(row=row, column=j+1, value=value)
                cell.font = normal_font
                cell.border = border
        
        # Recent problems
        row += 3
        ws_problems[f'A{row}'] = "Problèmes Récents"
        ws_problems[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        problem_headers = ['ID', 'Description', 'Localisation', 'Catégorie', 'Statut', 'Date de création']
        for j, header in enumerate(problem_headers):
            cell = ws_problems.cell(row=row, column=j+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for problem in report_data['problems']['recent']:
            row += 1
            problem_data = [
                str(problem['id']),
                problem['description'][:100] + '...' if len(problem['description']) > 100 else problem['description'],
                problem['location'] or 'N/A',
                problem['category__name'] or 'Sans catégorie',
                problem['status'],
                problem['created_at'].strftime('%d/%m/%Y %H:%M')
            ]
            for j, value in enumerate(problem_data):
                cell = ws_problems.cell(row=row, column=j+1, value=value)
                cell.font = normal_font
                cell.border = border
        
        # Auto-adjust column widths for problems sheet
        auto_adjust_columns(ws_problems)
    
    # Complaints sheet
    if 'complaints' in report_data:
        ws_complaints = wb.create_sheet("Analyse des Réclamations")
        
        # Title
        ws_complaints['A1'] = "ANALYSE DÉTAILLÉE DES RÉCLAMATIONS"
        ws_complaints['A1'].font = subheader_font
        ws_complaints['A1'].fill = subheader_fill
        ws_complaints.merge_cells('A1:E1')
        
        # Status distribution
        row = 3
        ws_complaints[f'A{row}'] = "Répartition par Statut"
        ws_complaints[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        status_headers = ['Statut', 'Nombre', 'Pourcentage']
        for j, header in enumerate(status_headers):
            cell = ws_complaints.cell(row=row, column=j+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        total_complaints = report_data['complaints']['total']
        for item in report_data['complaints']['by_status']:
            row += 1
            percentage = (item['count'] / max(total_complaints, 1)) * 100
            status_data = [item['status'], item['count'], f"{percentage:.1f}%"]
            for j, value in enumerate(status_data):
                cell = ws_complaints.cell(row=row, column=j+1, value=value)
                cell.font = normal_font
                cell.border = border
        
        # Recent complaints
        row += 3
        ws_complaints[f'A{row}'] = "Réclamations Récentes"
        ws_complaints[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        complaint_headers = ['ID', 'Sujet', 'Description', 'Statut', 'Date de création']
        for j, header in enumerate(complaint_headers):
            cell = ws_complaints.cell(row=row, column=j+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for complaint in report_data['complaints']['recent']:
            row += 1
            complaint_data = [
                str(complaint['id']),
                complaint['subject'],
                complaint['description'][:100] + '...' if len(complaint['description']) > 100 else complaint['description'],
                complaint['status'],
                complaint['created_at'].strftime('%d/%m/%Y %H:%M')
            ]
            for j, value in enumerate(complaint_data):
                cell = ws_complaints.cell(row=row, column=j+1, value=value)
                cell.font = normal_font
                cell.border = border
        
        # Auto-adjust column widths for complaints sheet
        auto_adjust_columns(ws_complaints)
    
    # Time Series sheet
    if report_data['time_series']['dates']:
        ws_time = wb.create_sheet("Évolution Temporelle")
        
        # Title
        ws_time['A1'] = "ÉVOLUTION TEMPORELLE DES SIGNALEMENTS"
        ws_time['A1'].font = subheader_font
        ws_time['A1'].fill = subheader_fill
        ws_time.merge_cells('A1:F1')
        
        # Headers
        row = 3
        time_headers = ['Date', 'Problèmes Total', 'Problèmes En Attente', 'Problèmes En Cours', 'Problèmes Résolus', 'Réclamations']
        for j, header in enumerate(time_headers):
            cell = ws_time.cell(row=row, column=j+1, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Data
        dates = report_data['time_series']['dates']
        problems_data = report_data['time_series']['problems']
        complaints_data = report_data['time_series']['complaints']
        
        for i, date_str in enumerate(dates):
            row += 1
            time_data = [
                datetime.datetime.strptime(date_str, '%Y-%m-%d').strftime('%d/%m/%Y'),
                problems_data['total'][i],
                problems_data['pending'][i],
                problems_data['in_progress'][i],
                problems_data['resolved'][i],
                complaints_data[i]
            ]
            for j, value in enumerate(time_data):
                cell = ws_time.cell(row=row, column=j+1, value=value)
                cell.font = normal_font
                cell.border = border
        
        # Auto-adjust column widths for time series sheet
        auto_adjust_columns(ws_time)
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_content = excel_buffer.getvalue()
    excel_buffer.close()
    
    # Create response
    response = HttpResponse(
        excel_content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"rapport_{report_data['municipality'].name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

# Complaint Management Views

@login_required
def complaint_list(request):
    """List all complaints for the municipality with filtering and pagination"""
    user_municipality = request.user.admin_profile.municipality

    search_term = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    complaints = Complaint.objects.filter(municipality=user_municipality)

    # Apply filters
    if search_term:
        complaints = complaints.filter(
            Q(subject__icontains=search_term) | 
            Q(description__icontains=search_term) |
            Q(citizen__full_name__icontains=search_term)
        )

    if status_filter:
        complaints = complaints.filter(status=status_filter)

    if date_from:
        try:
            date_from_parsed = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__date__gte=date_from_parsed)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_parsed = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            complaints = complaints.filter(created_at__date__lte=date_to_parsed)
        except ValueError:
            pass

    # Order by creation date (newest first)
    complaints = complaints.order_by('-created_at').select_related('citizen', 'municipality')

    # Pagination
    paginator = Paginator(complaints, ITEMS_PER_PAGE)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'complaints': page_obj,
        'search_term': search_term,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'status_choices': Complaint.STATUS_CHOICES,
        'navName': 'complaints',
    }

    return render(request, 'muni_admin/complaint_list.html', context)

@login_required
def complaint_detail(request, complaint_id):
    """View and manage a specific complaint"""
    complaint = get_object_or_404(
        Complaint.objects.select_related('citizen', 'municipality'), 
        pk=complaint_id, 
        municipality=request.user.admin_profile.municipality
    )
    
    # Fetch status change history for this complaint
    status_logs = StatusLog.objects.filter(
        record_type='COMPLAINT', 
        record_id=complaint.id
    ).order_by('-changed_at').select_related('changed_by')
    
    # Handle status update form submission
    if request.method == 'POST':
        old_status = complaint.status
        new_status = request.POST.get('status')
        comment = request.POST.get('comment', '')
        
        if new_status and new_status != old_status:
            # Update complaint status
            complaint.status = new_status
            complaint.comment = comment
            complaint.save()
            
            # Create status log entry
            StatusLog.objects.create(
                record_type='COMPLAINT',
                record_id=complaint.id,
                old_status=old_status,
                new_status=new_status,
                changed_by=request.user,
                changed_at=timezone.now(),
                comment=comment
            )
            
            # Create notification for the citizen
            Notification.objects.create(
                user=complaint.citizen.user,
                title=_("Mise à jour de votre réclamation"),
                message=_("Le statut de votre réclamation a été mis à jour à: {}").format(
                    dict(Complaint.STATUS_CHOICES)[new_status]
                ),
                type='COMPLAINT_UPDATE',
                related_id=str(complaint.id),
                related_type='COMPLAINT'
            )
            
            # Add success message
            from django.contrib import messages
            messages.success(request, _("Statut mis à jour avec succès."))
            
            # Redirect to avoid form resubmission
            return redirect('Muni_admin:complaint_detail', complaint_id=complaint_id)
    
    context = {
        'complaint': complaint,
        'status_logs': status_logs,
        'status_choices': Complaint.STATUS_CHOICES,
        'navName': 'complaints',
    }
    
    return render(request, 'muni_admin/complaint_detail.html', context)

from Citoyen.models import Citizen
from Citoyen.models import Citizen
from django.db.models import Q

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Count, Case, When, IntegerField


ITEMS_PER_PAGE = 12  # Adjust as needed

@login_required
def citizen_list(request):
    """List all citizens who have submitted problems in the municipality"""
    user_municipality = request.user.admin_profile.municipality
    
    search_term = request.GET.get('search', '')
    sort_by = request.GET.get('sort', 'full_name')
    
    # Filter citizens who have submitted problems in this municipality
    # and annotate with counts for each citizen
    citizens = Citizen.objects.filter(
         Q(problems__municipality=user_municipality) | 
        Q(complaints__municipality=user_municipality) 
        
    ).distinct().annotate(
        # Count problems for each citizen in this municipality
        problems_count=Count(
            'problems',
            filter=Q(problems__municipality=user_municipality),
            distinct=True
        ),
        # Count complaints for each citizen in this municipality
        complaints_count=Count(
            'complaints',
            filter=Q(complaints__municipality=user_municipality),
            distinct=True
        ),
        # Count resolved problems for each citizen in this municipality
        resolved_count=Count(
            'problems',
            filter=Q(
                problems__municipality=user_municipality,
                problems__status='RESOLVED'
            ),
            distinct=True
        )
    )
    
    if search_term:
        citizens = citizens.filter(
            Q(full_name__icontains=search_term) |
            Q(nni__icontains=search_term) |
            Q(user__phone_number__icontains=search_term) |
            Q(address__icontains=search_term)
        )
    
    # Apply sorting
    valid_sort_fields = {
        'full_name': 'full_name',
        '-full_name': '-full_name',
        'user__date_joined': 'user__date_joined',
        '-user__date_joined': '-user__date_joined',
        'nni': 'nni',
        'problems_count': 'problems_count',
        '-problems_count': '-problems_count',
        'complaints_count': 'complaints_count',
        '-complaints_count': '-complaints_count',
    }
    
    if sort_by in valid_sort_fields:
        citizens = citizens.order_by(valid_sort_fields[sort_by])
    else:
        citizens = citizens.order_by('full_name')
    
    # Select related to optimize queries
    citizens = citizens.select_related('user', 'municipality')
    
    # Calculate overall statistics for the header cards
    total_citizens = citizens.count()
    verified_count = citizens.filter(user__is_verified=True).count() if hasattr(citizens.first().user if citizens.first() else None, 'is_verified') else 0
    
    # Calculate total problems and complaints across all citizens
    total_problems = Problem.objects.filter(
        municipality=user_municipality,
        citizen__in=citizens
    ).count()
    
    total_complaints = Complaint.objects.filter(
        municipality=user_municipality,
        citizen__in=citizens
    ).count()
    
    # Calculate active reports (assuming pending/in_progress status)
    active_reports = Problem.objects.filter(
        municipality=user_municipality,
        citizen__in=citizens,
        status__in=['PENDING', 'IN_PROGRESS']
    ).count()
    
    # Calculate new citizens this month
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    current_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    new_this_month = citizens.filter(
        user__date_joined__gte=current_month
    ).count()
    
    # Calculate verification percentage
    verified_percentage = round((verified_count / total_citizens * 100) if total_citizens > 0 else 0, 1)
    
    # Pagination
    paginator = Paginator(citizens, ITEMS_PER_PAGE)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'citizens': page_obj,
        'search_term': search_term,
        'navName': 'citizens',
        # Overall statistics for header cards
        'verified_count': verified_count,
        'verified_percentage': verified_percentage,
        'active_reports': active_reports,
        'new_this_month': new_this_month,
        'total_problems': total_problems,
        'total_complaints': total_complaints,
        'current_sort': sort_by,
    }
    
    return render(request, 'muni_admin/citizen_list.html', context)

@login_required
def citizen_detail(request, citizen_id):
    """View detailed information about a citizen and their submissions"""
    citizen = get_object_or_404(
        Citizen.objects.select_related('user', 'municipality'), 
        pk=citizen_id, 
        municipality=request.user.admin_profile.municipality
    )
    
    # Get citizen's problems and complaints
    problems = Problem.objects.filter(citizen=citizen).order_by('-created_at')
    complaints = Complaint.objects.filter(citizen=citizen).order_by('-created_at')
    
    # Get statistics
    total_problems = problems.count()
    total_complaints = complaints.count()
    resolved_problems = problems.filter(status='RESOLVED').count()
    resolved_complaints = complaints.filter(status='RESOLVED').count()
    
    context = {
        'citizen': citizen,
        'problems': problems[:10],  # Show latest 10
        'complaints': complaints[:10],  # Show latest 10
        'total_problems': total_problems,
        'total_complaints': total_complaints,
        'resolved_problems': resolved_problems,
        'resolved_complaints': resolved_complaints,
        'navName': 'citizens',
    }
    
    return render(request, 'muni_admin/citizen_detail.html', context)

@login_required
def media_viewer(request, media_type, media_id):
    """View media files (photos, videos, documents) from problems and complaints"""
    user_municipality = request.user.admin_profile.municipality
    
    # Determine the model and field based on media_type and media_id
    if media_type == 'problem':
        try:
            problem = Problem.objects.get(id=media_id, municipality=user_municipality)
            media_files = {
                'photo': problem.photo,
                'video': problem.video,
                'voice_record': problem.voice_record,
                'document': problem.document,
            }
            context = {
                'object': problem,
                'object_type': 'problem',
                'media_files': media_files,
            }
        except Problem.DoesNotExist:
            return JsonResponse({'error': 'Problem not found'}, status=404)
    
    elif media_type == 'complaint':
        try:
            complaint = Complaint.objects.get(id=media_id, municipality=user_municipality)
            media_files = {
                'photo': complaint.photo,
                'video': complaint.video,
                'voice_record': complaint.voice_record,
                'evidence': complaint.evidence,
            }
            context = {
                'object': complaint,
                'object_type': 'complaint',
                'media_files': media_files,
            }
        except Complaint.DoesNotExist:
            return JsonResponse({'error': 'Complaint not found'}, status=404)
    
    else:
        return JsonResponse({'error': 'Invalid media type'}, status=400)
    
    return render(request, 'muni_admin/media_viewer.html', context)

@login_required
def bulk_actions(request):
    """Handle bulk actions on problems and complaints"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    action = request.POST.get('action')
    item_type = request.POST.get('type')  # 'problem' or 'complaint'
    item_ids = request.POST.getlist('items')
    
    if not action or not item_type or not item_ids:
        return JsonResponse({'error': 'Missing required parameters'}, status=400)
    
    user_municipality = request.user.admin_profile.municipality
    
    try:
        if item_type == 'problem':
            items = Problem.objects.filter(id__in=item_ids, municipality=user_municipality)
        elif item_type == 'complaint':
            items = Complaint.objects.filter(id__in=item_ids, municipality=user_municipality)
        else:
            return JsonResponse({'error': 'Invalid item type'}, status=400)
        
        if action == 'update_status':
            new_status = request.POST.get('new_status')
            if not new_status:
                return JsonResponse({'error': 'New status required'}, status=400)
            
            updated_count = 0
            for item in items:
                old_status = item.status
                if old_status != new_status:
                    item.status = new_status
                    item.save()
                    
                    # Create status log
                    StatusLog.objects.create(
                        record_type=item_type.upper(),
                        record_id=item.id,
                        old_status=old_status,
                        new_status=new_status,
                        changed_by=request.user,
                        changed_at=timezone.now(),
                        comment=f"Mise à jour en lot par {request.user.username}"
                    )
                    
                    # Create notification
                    Notification.objects.create(
                        user=item.citizen.user,
                        title=f"Mise à jour de votre {item_type}",
                        message=f"Le statut a été mis à jour à: {dict(item.STATUS_CHOICES)[new_status]}",
                        type=f"{item_type.upper()}_UPDATE",
                        related_id=str(item.id),
                        related_type=item_type.upper()
                    )
                    
                    updated_count += 1
            
            return JsonResponse({
                'success': True,
                'message': f'{updated_count} éléments mis à jour avec succès'
            })
        
        elif action == 'export':
            # Export selected items to Excel
            return _export_items_to_excel(items, item_type)
        
        else:
            return JsonResponse({'error': 'Unknown action'}, status=400)
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def _export_items_to_excel(items, item_type):
    """Export items to Excel format"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return JsonResponse({'error': 'Excel export not available'}, status=500)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{item_type.title()}s"
    
    # Headers
    if item_type == 'problem':
        headers = ['ID', 'Description', 'Localisation', 'Catégorie', 'Statut', 'Citoyen', 'Date de création']
        ws.append(headers)
        
        # Data
        for problem in items:
            ws.append([
                str(problem.id),
                problem.description,
                problem.location or 'N/A',
                problem.category.name if problem.category else 'Sans catégorie',
                problem.get_status_display(),
                problem.citizen.full_name,
                problem.created_at.strftime('%d/%m/%Y %H:%M')
            ])
    
    elif item_type == 'complaint':
        headers = ['ID', 'Sujet', 'Description', 'Statut', 'Citoyen', 'Date de création']
        ws.append(headers)
        
        # Data
        for complaint in items:
            ws.append([
                str(complaint.id),
                complaint.subject,
                complaint.description,
                complaint.get_status_display(),
                complaint.citizen.full_name,
                complaint.created_at.strftime('%d/%m/%Y %H:%M')
            ])
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"{item_type}s_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# muni_admin/views.py
# muni_admin/views.py
# Add these imports to your existing views.py file

import json
import logging
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.core.serializers.json import DjangoJSONEncoder
from django.utils.decorators import method_decorator
from django.views import View
from django.shortcuts import render
from django.utils import timezone
from django.middleware.csrf import get_token
from .chatbot_logic import get_chatbot_response, get_quick_stats, suggest_chatbot_questions, detect_language

logger = logging.getLogger(__name__)

@login_required
def chatbot_page(request):
    """Renders the HTML page for the chatbot interface."""
    context = {
        'navName': 'chatbot',
        'municipality_id': request.user.admin_profile.municipality.id if hasattr(request.user, 'admin_profile') and request.user.admin_profile.municipality else None,
        'quick_stats': get_quick_stats(request.user.admin_profile.municipality.id) if hasattr(request.user, 'admin_profile') and request.user.admin_profile.municipality else None,
        'suggested_questions': suggest_chatbot_questions('fr'),  # Default to French, can be made dynamic
        'csrf_token': get_token(request)  # Add CSRF token to context
    }
    return render(request, 'muni_admin/chatbot.html', context)

@method_decorator(csrf_exempt, name='dispatch')  # Exempt CSRF for API
@method_decorator(login_required, name='dispatch')
class ChatbotAPIView(View):
    """
    API endpoint for chatbot interactions with proper error handling and security
    """
    
    def post(self, request):
        try:
            # Parse request data
            data = json.loads(request.body)
            user_message = data.get('message', '').strip()
            
            if not user_message:
                return JsonResponse({
                    'error': 'Message cannot be empty',
                    'success': False
                }, status=400)
            
            # Get user's municipality
            if not hasattr(request.user, 'admin_profile') or not request.user.admin_profile.municipality:
                return JsonResponse({
                    'error': 'User not associated with a municipality',
                    'success': False
                }, status=403)
            
            municipality_id = request.user.admin_profile.municipality.id
            
            # Log the interaction
            logger.info(f"Chatbot interaction - User: {request.user.username}, Municipality: {municipality_id}, Message: {user_message}")
            
            # Get chatbot response
            bot_response = get_chatbot_response(user_message, municipality_id)
            
            # Detect language for suggestions
            user_language = detect_language(user_message)
            suggested_questions = suggest_chatbot_questions('fr' if user_language == 'fr' else 'en')
            
            return JsonResponse({
                'response': bot_response,
                'success': True,
                'suggested_questions': suggested_questions,
                'timestamp': timezone.now().isoformat()
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'error': 'Invalid JSON data',
                'success': False
            }, status=400)
        
        except Exception as e:
            logger.error(f"Chatbot API error: {e}", exc_info=True)
            return JsonResponse({
                'error': 'Internal server error',
                'success': False
            }, status=500)
    
    def get(self, request):
        """Get chatbot status and suggestions"""
        try:
            if not hasattr(request.user, 'admin_profile') or not request.user.admin_profile.municipality:
                return JsonResponse({
                    'error': 'User not associated with a municipality',
                    'success': False
                }, status=403)
            
            municipality_id = request.user.admin_profile.municipality.id
            quick_stats = get_quick_stats(municipality_id)
            
            return JsonResponse({
                'success': True,
                'quick_stats': quick_stats,
                'suggested_questions': {
                    'fr': suggest_chatbot_questions('fr'),
                    'en': suggest_chatbot_questions('en')
                },
                'municipality_id': municipality_id
            })
            
        except Exception as e:
            logger.error(f"Chatbot status error: {e}", exc_info=True)
            return JsonResponse({
                'error': 'Internal server error',
                'success': False
            }, status=500)

# Alternative function-based view (if you prefer)
@login_required
@csrf_exempt  # Exempt CSRF for API endpoint
@require_http_methods(["POST"])
def chatbot_api_view(request):
    """
    Function-based API endpoint for chatbot interactions
    """
    try:
        data = json.loads(request.body)
        user_message = data.get('message', '').strip()
        
        if not user_message:
            return JsonResponse({
                'error': 'Message is required',
                'success': False
            }, status=400)
        
        # Security check - ensure user has admin profile and municipality
        if not hasattr(request.user, 'admin_profile') or not request.user.admin_profile.municipality:
            return JsonResponse({
                'error': 'Unauthorized - No municipality access',
                'success': False
            }, status=403)
        
        municipality_id = request.user.admin_profile.municipality.id
        
        # Get chatbot response
        bot_response = get_chatbot_response(user_message, municipality_id)
        
        return JsonResponse({
            'response': bot_response,
            'success': True,
            'timestamp': timezone.now().isoformat()
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'error': 'Invalid JSON format',
            'success': False
        }, status=400)
    
    except Exception as e:
        logger.error(f"Chatbot API error: {e}", exc_info=True)
        return JsonResponse({
            'error': 'Server error occurred',
            'success': False
        }, status=500)

