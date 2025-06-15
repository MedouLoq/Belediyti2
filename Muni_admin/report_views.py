# Report Generation Views for Muni_admin

from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils.translation import gettext as _
from django.db.models import Q, Count, Avg
from django.utils import timezone
from django.template.loader import render_to_string
from django.conf import settings
import datetime
import json
import os
import tempfile
import subprocess
from collections import defaultdict

from Citoyen.models import Problem, Complaint, Category, StatusLog, Citizen, Municipality

@login_required
def report_dashboard(request):
    """Dashboard for report generation with various options"""
    municipality = request.user.admin_profile.municipality
    
    # Get available date ranges
    earliest_problem = Problem.objects.filter(municipality=municipality).order_by('created_at').first()
    earliest_complaint = Complaint.objects.filter(municipality=municipality).order_by('created_at').first()
    
    earliest_date = None
    if earliest_problem and earliest_complaint:
        earliest_date = min(earliest_problem.created_at, earliest_complaint.created_at).date()
    elif earliest_problem:
        earliest_date = earliest_problem.created_at.date()
    elif earliest_complaint:
        earliest_date = earliest_complaint.created_at.date()
    
    context = {
        'municipality': municipality,
        'earliest_date': earliest_date,
        'categories': Category.objects.all(),
        'navName': 'reports',
    }
    
    return render(request, 'muni_admin/reports/dashboard.html', context)

@login_required
def generate_report(request):
    """Generate a comprehensive report based on selected parameters"""
    if request.method != 'POST':
        return redirect('Muni_admin:report_dashboard')
    
    municipality = request.user.admin_profile.municipality
    
    # Get report parameters
    report_type = request.POST.get('report_type', 'comprehensive')
    date_from = request.POST.get('date_from')
    date_to = request.POST.get('date_to')
    include_problems = request.POST.get('include_problems') == 'on'
    include_complaints = request.POST.get('include_complaints') == 'on'
    include_charts = request.POST.get('include_charts') == 'on'
    include_maps = request.POST.get('include_maps') == 'on'
    categories = request.POST.getlist('categories')
    format_type = request.POST.get('format', 'pdf')
    
    # Parse dates
    try:
        if date_from:
            date_from = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
        if date_to:
            date_to = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)
    
    # Generate report data
    report_data = _generate_report_data(
        municipality, date_from, date_to, include_problems, 
        include_complaints, categories
    )
    
    # Generate report based on format
    if format_type == 'pdf':
        return _generate_pdf_report(request, report_data, include_charts, include_maps)
    elif format_type == 'excel':
        return _generate_excel_report(report_data)
    else:
        return JsonResponse({'error': 'Unsupported format'}, status=400)

def _generate_report_data(municipality, date_from, date_to, include_problems, include_complaints, categories):
    """Generate comprehensive report data"""
    data = {
        'municipality': municipality,
        'date_from': date_from,
        'date_to': date_to,
        'generated_at': timezone.now(),
        'generated_by': municipality.name,
    }
    
    # Base querysets
    problems_qs = Problem.objects.filter(municipality=municipality)
    complaints_qs = Complaint.objects.filter(municipality=municipality)
    
    # Apply date filters
    if date_from:
        problems_qs = problems_qs.filter(created_at__date__gte=date_from)
        complaints_qs = complaints_qs.filter(created_at__date__gte=date_from)
    
    if date_to:
        problems_qs = problems_qs.filter(created_at__date__lte=date_to)
        complaints_qs = complaints_qs.filter(created_at__date__lte=date_to)
    
    # Apply category filters
    if categories:
        problems_qs = problems_qs.filter(category_id__in=categories)
    
    # Problems data
    if include_problems:
        data['problems'] = {
            'total': problems_qs.count(),
            'by_status': list(problems_qs.values('status').annotate(count=Count('status'))),
            'by_category': list(problems_qs.values('category__name').annotate(count=Count('category'))),
            'recent': list(problems_qs.order_by('-created_at')[:10].values(
                'id', 'description', 'location', 'status', 'created_at', 'category__name'
            )),
            'resolution_stats': _calculate_resolution_stats(problems_qs),
        }
    
    # Complaints data
    if include_complaints:
        data['complaints'] = {
            'total': complaints_qs.count(),
            'by_status': list(complaints_qs.values('status').annotate(count=Count('status'))),
            'recent': list(complaints_qs.order_by('-created_at')[:10].values(
                'id', 'subject', 'description', 'status', 'created_at'
            )),
        }
    
    # Time series data
    data['time_series'] = _generate_time_series_data(problems_qs, complaints_qs, date_from, date_to)
    
    # Summary statistics
    data['summary'] = {
        'total_issues': (problems_qs.count() if include_problems else 0) + (complaints_qs.count() if include_complaints else 0),
        'resolved_issues': (problems_qs.filter(status='RESOLVED').count() if include_problems else 0) + 
                          (complaints_qs.filter(status='RESOLVED').count() if include_complaints else 0),
        'pending_issues': (problems_qs.filter(status='PENDING').count() if include_problems else 0) + 
                         (complaints_qs.filter(status='PENDING').count() if include_complaints else 0),
    }
    
    if data['summary']['total_issues'] > 0:
        data['summary']['resolution_rate'] = round(
            (data['summary']['resolved_issues'] / data['summary']['total_issues']) * 100, 2
        )
    else:
        data['summary']['resolution_rate'] = 0
    
    return data

def _calculate_resolution_stats(problems_qs):
    """Calculate resolution statistics for problems"""
    resolved_problems = problems_qs.filter(status='RESOLVED')
    
    if not resolved_problems.exists():
        return {'avg_resolution_time': 0, 'by_category': []}
    
    # Get resolution times from status logs
    resolution_times = []
    category_times = defaultdict(list)
    
    for problem in resolved_problems:
        log_entry = StatusLog.objects.filter(
            record_type='PROBLEM',
            record_id=problem.id,
            new_status='RESOLVED'
        ).first()
        
        if log_entry:
            resolution_time = (log_entry.changed_at.date() - problem.created_at.date()).days
            resolution_times.append(resolution_time)
            
            category_name = problem.category.name if problem.category else 'Sans catégorie'
            category_times[category_name].append(resolution_time)
    
    avg_resolution_time = sum(resolution_times) / len(resolution_times) if resolution_times else 0
    
    by_category = []
    for category, times in category_times.items():
        by_category.append({
            'category': category,
            'avg_days': sum(times) / len(times),
            'count': len(times)
        })
    
    return {
        'avg_resolution_time': round(avg_resolution_time, 1),
        'by_category': sorted(by_category, key=lambda x: x['avg_days'])
    }

def _generate_time_series_data(problems_qs, complaints_qs, date_from, date_to):
    """Generate time series data for charts"""
    if not date_from:
        date_from = timezone.now().date() - datetime.timedelta(days=30)
    if not date_to:
        date_to = timezone.now().date()
    
    # Generate date range
    date_range = []
    current_date = date_from
    while current_date <= date_to:
        date_range.append(current_date)
        current_date += datetime.timedelta(days=1)
    
    # Problems by day
    problems_by_day = problems_qs.extra(
        select={'day': 'date(created_at)'}
    ).values('day').annotate(count=Count('id'))
    
    problems_data = {item['day'].strftime('%Y-%m-%d'): item['count'] for item in problems_by_day}
    
    # Complaints by day
    complaints_by_day = complaints_qs.extra(
        select={'day': 'date(created_at)'}
    ).values('day').annotate(count=Count('id'))
    
    complaints_data = {item['day'].strftime('%Y-%m-%d'): item['count'] for item in complaints_by_day}
    
    # Build series data
    problems_series = [problems_data.get(d.strftime('%Y-%m-%d'), 0) for d in date_range]
    complaints_series = [complaints_data.get(d.strftime('%Y-%m-%d'), 0) for d in date_range]
    
    return {
        'dates': [d.strftime('%Y-%m-%d') for d in date_range],
        'problems': problems_series,
        'complaints': complaints_series
    }

def _generate_pdf_report(request, report_data, include_charts, include_maps):
    """Generate PDF report using markdown and manus-md-to-pdf"""
    
    # Render markdown content
    markdown_content = render_to_string('muni_admin/reports/report_template.md', {
        'data': report_data,
        'include_charts': include_charts,
        'include_maps': include_maps,
    })
    
    # Create temporary files
    with tempfile.NamedTemporaryFile(mode='w', suffix='.md', delete=False) as md_file:
        md_file.write(markdown_content)
        md_file_path = md_file.name
    
    pdf_file_path = md_file_path.replace('.md', '.pdf')
    
    try:
        # Convert markdown to PDF using manus utility
        result = subprocess.run([
            'manus-md-to-pdf', md_file_path, pdf_file_path
        ], capture_output=True, text=True)
        
        if result.returncode != 0:
            return JsonResponse({'error': f'PDF generation failed: {result.stderr}'}, status=500)
        
        # Read PDF file
        with open(pdf_file_path, 'rb') as pdf_file:
            pdf_content = pdf_file.read()
        
        # Create response
        response = HttpResponse(pdf_content, content_type='application/pdf')
        filename = f"rapport_{report_data['municipality'].name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    finally:
        # Clean up temporary files
        try:
            os.unlink(md_file_path)
            if os.path.exists(pdf_file_path):
                os.unlink(pdf_file_path)
        except OSError:
            pass

def _generate_excel_report(report_data):
    """Generate Excel report"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        return JsonResponse({'error': 'Excel generation not available'}, status=500)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Résumé"
    
    # Add summary data
    ws_summary['A1'] = "Rapport Municipal"
    ws_summary['A1'].font = Font(size=16, bold=True)
    
    ws_summary['A3'] = "Municipalité:"
    ws_summary['B3'] = report_data['municipality'].name
    
    ws_summary['A4'] = "Période:"
    period_text = f"Du {report_data['date_from']} au {report_data['date_to']}" if report_data['date_from'] and report_data['date_to'] else "Toutes les données"
    ws_summary['B4'] = period_text
    
    ws_summary['A5'] = "Généré le:"
    ws_summary['B5'] = report_data['generated_at'].strftime('%d/%m/%Y %H:%M')
    
    # Summary statistics
    row = 7
    ws_summary[f'A{row}'] = "Statistiques Générales"
    ws_summary[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws_summary[f'A{row}'] = "Total des signalements:"
    ws_summary[f'B{row}'] = report_data['summary']['total_issues']
    
    row += 1
    ws_summary[f'A{row}'] = "Signalements résolus:"
    ws_summary[f'B{row}'] = report_data['summary']['resolved_issues']
    
    row += 1
    ws_summary[f'A{row}'] = "Taux de résolution:"
    ws_summary[f'B{row}'] = f"{report_data['summary']['resolution_rate']}%"
    
    # Problems sheet
    if 'problems' in report_data:
        ws_problems = wb.create_sheet("Problèmes")
        
        # Headers
        headers = ['ID', 'Description', 'Localisation', 'Catégorie', 'Statut', 'Date de création']
        for col, header in enumerate(headers, 1):
            cell = ws_problems.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, problem in enumerate(report_data['problems']['recent'], 2):
            ws_problems.cell(row=row, column=1, value=str(problem['id']))
            ws_problems.cell(row=row, column=2, value=problem['description'][:50] + '...' if len(problem['description']) > 50 else problem['description'])
            ws_problems.cell(row=row, column=3, value=problem['location'] or 'N/A')
            ws_problems.cell(row=row, column=4, value=problem['category__name'] or 'Sans catégorie')
            ws_problems.cell(row=row, column=5, value=problem['status'])
            ws_problems.cell(row=row, column=6, value=problem['created_at'].strftime('%d/%m/%Y %H:%M'))
    
    # Complaints sheet
    if 'complaints' in report_data:
        ws_complaints = wb.create_sheet("Réclamations")
        
        # Headers
        headers = ['ID', 'Sujet', 'Description', 'Statut', 'Date de création']
        for col, header in enumerate(headers, 1):
            cell = ws_complaints.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, complaint in enumerate(report_data['complaints']['recent'], 2):
            ws_complaints.cell(row=row, column=1, value=str(complaint['id']))
            ws_complaints.cell(row=row, column=2, value=complaint['subject'])
            ws_complaints.cell(row=row, column=3, value=complaint['description'][:50] + '...' if len(complaint['description']) > 50 else complaint['description'])
            ws_complaints.cell(row=row, column=4, value=complaint['status'])
            ws_complaints.cell(row=row, column=5, value=complaint['created_at'].strftime('%d/%m/%Y %H:%M'))
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"rapport_{report_data['municipality'].name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def scheduled_reports(request):
    """Manage scheduled reports"""
    # This would integrate with a task scheduler like Celery
    # For now, just show the interface
    context = {
        'navName': 'reports',
    }
    return render(request, 'muni_admin/reports/scheduled.html', context)

@login_required
def report_templates(request):
    """Manage report templates"""
    context = {
        'navName': 'reports',
    }
    return render(request, 'muni_admin/reports/templates.html', context)

